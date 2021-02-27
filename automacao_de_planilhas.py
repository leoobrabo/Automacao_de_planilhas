import openpyxl

planilha = openpyxl.Workbook()


def criar_pagina(nome):
    planilha.create_sheet(nome)
    print(f'Planilha {nome} criada com sucesso!')


def criar_cabecalho():
    nome_cabecalho = input('Digite o Nome do cabeçalho: ')
    return lista_cabecalho.append(nome_cabecalho)


def selecionar_pagina():
    escolha_pagina = input(
        (f'Escolha uma dessas paginas: {planilha.sheetnames}: '))
    pagina_selecionada = planilha[escolha_pagina]
    return pagina_selecionada


print('Bem Vindo ao gerador de Planilhas!!!')
print('Para começar vamos criar uma nova página dentro de uma planilha! ')
nome_planilha = input('Digite o nome da pagina a ser criada: ')
criar_pagina(nome_planilha)

nova_pagina = True

while nova_pagina:
    escolha = input('Deseja Criar mais uma pagina nesta planilha (s/n): ')
    if escolha == 'n' or escolha == 'não' or escolha == 'nao':
        nova_pagina = False
    elif escolha == 's' or escolha == 'sim':
        nome_planilha = input('Digite o nome da pagina a ser criada: ')
        criar_pagina(nome_planilha)
        nova_pagina = True
    else:
        print('Opção Invalida!!! ')

print('Escolha a pagina em que deseja trabalhar!')
selecao = selecionar_pagina()
lista_cabecalho = []
adcionar_cabecalho = True

while adcionar_cabecalho:
    escolha = input('Deseja criar cabecalho (s/n): ')
    if escolha == 's':
        criar_cabecalho()
    elif escolha == 'n':
        selecao.append(lista_cabecalho)
        adcionar_cabecalho = False
    else:
        print('Opção invalida!')

adicionar_nova_linha = True

while adicionar_nova_linha:
    dados = input(
        'Digite os dados a serem adicionados, separados por virgula:  ')
    selecao.append(dados.split(','))
    resposta = input('Adicionar mais uma linha?(s/n) ')
    if resposta == 's':
        adicionar_nova_linha = True
    else:
        adicionar_nova_linha = False

nome_planilha = input('Digite o nome da planilha a ser salva: ')
planilha.save(f'{nome_planilha}.xlsx')
print('Finalizando programa')
